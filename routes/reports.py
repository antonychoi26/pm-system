"""Reports routes - Weekly report generation + Excel import"""
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, send_file, abort)
from flask_login import login_required, current_user
from models import db, Task, TaskLog, TaskStatus, TaskCategory, Estate, User
from datetime import datetime, date, timedelta
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)


def get_accessible_estate_ids():
    if current_user.is_admin:
        return [e.id for e in Estate.query.filter_by(is_active=True).all()]
    return current_user.assigned_estate_ids


# ── Report Generator ─────────────────────────────────────────────────────────
@reports_bp.route('/')
@login_required
def index():
    estates = _get_estates()
    statuses = TaskStatus.query.filter_by(is_active=True).order_by(TaskStatus.sort_order).all()
    categories = TaskCategory.query.filter_by(is_active=True).order_by(TaskCategory.sort_order).all()

    today = date.today()
    # Default: from Monday to today
    week_start = today - timedelta(days=today.weekday())

    return render_template('reports/index.html',
        estates=estates, statuses=statuses, categories=categories,
        default_from=week_start.strftime('%Y-%m-%d'),
        default_to=today.strftime('%Y-%m-%d'),
    )


@reports_bp.route('/generate')
@login_required
def generate():
    """Generate report with user-specified filters."""
    estate_ids_all = get_accessible_estate_ids()

    # Filters
    f_estates    = request.args.getlist('estate_ids', type=int)
    f_status     = request.args.getlist('status_ids', type=int)
    f_date_from  = request.args.get('date_from', '')
    f_date_to    = request.args.get('date_to', '')
    f_include_terminal = bool(request.args.get('include_terminal'))
    report_title = request.args.get('title', '屋苑每週事項跟進報告')

    # Scope to accessible estates
    if f_estates:
        scope_ids = [eid for eid in f_estates if eid in estate_ids_all]
    else:
        scope_ids = estate_ids_all

    # Build task query
    q = Task.query.filter(Task.estate_id.in_(scope_ids), Task.is_active == True)

    if f_status:
        q = q.filter(Task.status_id.in_(f_status))
    elif not f_include_terminal:
        terminal_ids = [s.id for s in TaskStatus.query.filter_by(is_terminal=True).all()]
        q = q.filter(~Task.status_id.in_(terminal_ids))

    # Date filter on log entries (tasks with logs in date range)
    if f_date_from or f_date_to:
        log_task_ids = _get_task_ids_with_logs_in_range(f_date_from, f_date_to)
        # Also include tasks created in range
        create_q = Task.query.filter(Task.estate_id.in_(scope_ids), Task.is_active == True)
        if f_date_from:
            try:
                d = datetime.strptime(f_date_from, '%Y-%m-%d')
                create_q = create_q.filter(Task.created_at >= d)
            except ValueError:
                pass
        if f_date_to:
            try:
                d = datetime.strptime(f_date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
                create_q = create_q.filter(Task.created_at <= d)
            except ValueError:
                pass
        create_ids = [t.id for t in create_q.all()]
        all_ids = list(set(log_task_ids + create_ids))
        q = q.filter(Task.id.in_(all_ids))

    tasks = q.order_by(Task.estate_id, Task.created_at).all()

    # Group by estate
    estate_groups = {}
    for task in tasks:
        est = task.estate
        if est not in estate_groups:
            estate_groups[est] = []

        # Filter logs in date range
        logs_q = task.logs.order_by(TaskLog.log_date.asc(), TaskLog.created_at.asc())
        if f_date_from or f_date_to:
            if f_date_from:
                try:
                    logs_q = logs_q.filter(TaskLog.log_date >= datetime.strptime(f_date_from, '%Y-%m-%d').date())
                except ValueError:
                    pass
            if f_date_to:
                try:
                    logs_q = logs_q.filter(TaskLog.log_date <= datetime.strptime(f_date_to, '%Y-%m-%d').date())
                except ValueError:
                    pass
        logs = logs_q.all()

        # Show all logs if no date filter
        if not f_date_from and not f_date_to:
            logs = task.logs.order_by(TaskLog.log_date.asc()).all()

        estate_groups[est].append({'task': task, 'logs': logs})

    # Sort estates by code
    sorted_groups = sorted(estate_groups.items(), key=lambda x: x[0].code)

    format_type = request.args.get('format', 'html')
    if format_type == 'excel':
        return _export_excel(sorted_groups, report_title, f_date_from, f_date_to)

    # Compute week number
    try:
        ref_date = datetime.strptime(f_date_to, '%Y-%m-%d') if f_date_to else date.today()
    except ValueError:
        ref_date = date.today()
    if isinstance(ref_date, datetime):
        ref_date = ref_date.date()
    week_num = ref_date.isocalendar()[1]

    return render_template('reports/view.html',
        sorted_groups=sorted_groups,
        report_title=report_title,
        f_date_from=f_date_from,
        f_date_to=f_date_to,
        week_num=week_num,
        ref_date=ref_date,
    )


def _get_task_ids_with_logs_in_range(date_from, date_to):
    q = TaskLog.query
    if date_from:
        try:
            q = q.filter(TaskLog.log_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(TaskLog.log_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    return [l.task_id for l in q.all()]


def _export_excel(sorted_groups, title, date_from, date_to):
    """Generate Excel report matching original format."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Styles
    hdr_font  = Font(bold=True, size=11)
    title_font= Font(bold=True, size=14)
    hdr_fill  = PatternFill('solid', fgColor='366092')
    hdr_font_w= Font(bold=True, color='FFFFFF', size=10)
    sub_fill  = PatternFill('solid', fgColor='DCE6F1')
    wrap_align= Alignment(wrap_text=True, vertical='top')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_COLORS = {
        'info':      'D0ECF7',
        'warning':   'FFF3CD',
        'success':   'D4EDDA',
        'secondary': 'E2E3E5',
        'danger':    'F8D7DA',
        'primary':   'CCE5FF',
    }

    for estate, task_items in sorted_groups:
        safe_name = estate.code[:31]
        ws = wb.create_sheet(title=safe_name)

        # Header
        ws.merge_cells('A1:I1')
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        ws.merge_cells('A2:C2')
        ws['A2'] = f'屋苑：{estate.display_name}'
        ws['A2'].font = Font(bold=True, size=11)

        ws.merge_cells('G2:H2')
        date_label = ''
        if date_from and date_to:
            date_label = f'{date_from} - {date_to}'
        ws['G2'] = date_label
        ws['G2'].alignment = Alignment(horizontal='right')
        ws['I2'] = f'報告日期: {date.today().strftime("%Y-%m-%d")}'

        # Column headers
        headers = ['事項編號', '事項種類', '事項狀況', '事項新增日期',
                   '事項內容', '跟進日期', '事項跟進記錄', '預計完成日期', '跟進人']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = hdr_font_w
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border

        # Column widths
        col_widths = [14, 16, 10, 14, 30, 12, 45, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 5
        for item in task_items:
            task = item['task']
            logs = item['logs']

            status_color = STATUS_COLORS.get(
                task.status_obj.color if task.status_obj else 'secondary', 'FFFFFF')
            task_fill = PatternFill('solid', fgColor=status_color)

            if not logs:
                # Task row without logs
                ws.cell(row=row, column=1, value=task.task_number)
                ws.cell(row=row, column=2, value=task.category.name_zh if task.category else '')
                ws.cell(row=row, column=3, value=task.status_obj.name_zh if task.status_obj else '')
                ws.cell(row=row, column=4, value=task.created_at.strftime('%d/%m/%Y') if task.created_at else '')
                ws.cell(row=row, column=5, value=task.title)
                ws.cell(row=row, column=6, value='')
                ws.cell(row=row, column=7, value='')
                ws.cell(row=row, column=8, value=task.expected_done.strftime('%d/%m/%Y') if task.expected_done else '')
                ws.cell(row=row, column=9, value=task.assignee.display_name if task.assignee else '')
                for col in range(1, 10):
                    c = ws.cell(row=row, column=col)
                    c.fill = task_fill
                    c.border = border
                    c.alignment = wrap_align
                row += 1
            else:
                for i, log in enumerate(logs):
                    if i == 0:
                        ws.cell(row=row, column=1, value=task.task_number)
                        ws.cell(row=row, column=2, value=task.category.name_zh if task.category else '')
                        ws.cell(row=row, column=3, value=task.status_obj.name_zh if task.status_obj else '')
                        ws.cell(row=row, column=4, value=task.created_at.strftime('%d/%m/%Y') if task.created_at else '')
                        ws.cell(row=row, column=5, value=task.title)
                        ws.cell(row=row, column=8, value=task.expected_done.strftime('%d/%m/%Y') if task.expected_done else '')
                        ws.cell(row=row, column=9, value=task.assignee.display_name if task.assignee else '')
                    ws.cell(row=row, column=6, value=log.log_date.strftime('%d/%m/%Y') if log.log_date else '')
                    ws.cell(row=row, column=7, value=log.content)
                    for col in range(1, 10):
                        c = ws.cell(row=row, column=col)
                        c.fill = task_fill if i == 0 else PatternFill('solid', fgColor='F5F5F5')
                        c.border = border
                        c.alignment = wrap_align
                    row += 1

        ws.freeze_panes = 'A5'

    # Master list sheet
    _add_master_sheet(wb, sorted_groups, title)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"Working_List_Report_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname)


def _add_master_sheet(wb, sorted_groups, title):
    ws = wb.create_sheet(title='Master List', index=0)
    hdr_fill  = PatternFill('solid', fgColor='366092')
    hdr_font_w= Font(bold=True, color='FFFFFF', size=10)
    title_font= Font(bold=True, size=14)
    wrap_align= Alignment(wrap_text=True, vertical='top')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    headers = ['屋苑', '事項編號', '事項種類', '事項狀況', '事項新增日期',
               '事項內容', '跟進日期', '事項跟進記錄', '預計完成日期', '跟進人']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = hdr_font_w
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    col_widths = [18, 14, 16, 10, 14, 30, 12, 45, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    STATUS_COLORS = {'info': 'D0ECF7', 'warning': 'FFF3CD', 'success': 'D4EDDA',
                     'secondary': 'E2E3E5', 'danger': 'F8D7DA', 'primary': 'CCE5FF'}

    for estate, task_items in sorted_groups:
        for item in task_items:
            task = item['task']
            logs = item['logs']
            sc   = STATUS_COLORS.get(task.status_obj.color if task.status_obj else 'secondary', 'FFFFFF')
            fill = PatternFill('solid', fgColor=sc)

            if not logs:
                data = [estate.display_name, task.task_number,
                        task.category.name_zh if task.category else '',
                        task.status_obj.name_zh if task.status_obj else '',
                        task.created_at.strftime('%d/%m/%Y') if task.created_at else '',
                        task.title, '', '', 
                        task.expected_done.strftime('%d/%m/%Y') if task.expected_done else '',
                        task.assignee.display_name if task.assignee else '']
                for col_idx, val in enumerate(data, 1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.fill = fill; c.border = border; c.alignment = wrap_align
                row += 1
            else:
                for i, log in enumerate(logs):
                    if i == 0:
                        ws.cell(row=row, column=1, value=estate.display_name)
                        ws.cell(row=row, column=2, value=task.task_number)
                        ws.cell(row=row, column=3, value=task.category.name_zh if task.category else '')
                        ws.cell(row=row, column=4, value=task.status_obj.name_zh if task.status_obj else '')
                        ws.cell(row=row, column=5, value=task.created_at.strftime('%d/%m/%Y') if task.created_at else '')
                        ws.cell(row=row, column=6, value=task.title)
                        ws.cell(row=row, column=9, value=task.expected_done.strftime('%d/%m/%Y') if task.expected_done else '')
                        ws.cell(row=row, column=10, value=task.assignee.display_name if task.assignee else '')
                    ws.cell(row=row, column=7, value=log.log_date.strftime('%d/%m/%Y') if log.log_date else '')
                    ws.cell(row=row, column=8, value=log.content)
                    rfill = fill if i == 0 else PatternFill('solid', fgColor='F5F5F5')
                    for col in range(1, 11):
                        c = ws.cell(row=row, column=col)
                        c.fill = rfill; c.border = border; c.alignment = wrap_align
                    row += 1

    ws.freeze_panes = 'A4'


# ── Import Excel ─────────────────────────────────────────────────────────────
@reports_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    if not current_user.is_manager:
        abort(403)

    result = None
    error  = None

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith('.xlsx'):
            error = '請上載 .xlsx 格式的 Excel 檔案。'
        else:
            try:
                result = _do_import(file)
                flash(f'匯入完成！新增 {result["created"]} 個事項，{result["logs_added"]} 條跟進記錄。', 'success')
            except Exception as e:
                error = f'匯入失敗：{str(e)}'

    return render_template('reports/import.html', result=result, error=error)


def _do_import(file_obj):
    """Import from Working_List_DataBase.xlsx format."""
    import re
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb['DataBase'] if 'DataBase' in wb.sheetnames else wb.active

    # Build lookup maps
    estate_map   = {e.code: e for e in Estate.query.all()}
    cat_map      = {c.name_zh: c for c in TaskCategory.query.all()}
    status_map   = {s.name_zh: s for s in TaskStatus.query.all()}
    user_map     = {u.display_name: u for u in User.query.all()}

    # Status normalisation
    status_aliases = {
        '完成': '完成', '已完成': '已完成', '跟進中': '跟進中', '新增': '新增',
    }

    created_count = 0
    logs_added    = 0
    task_cache    = {}  # task_number -> Task

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue

        task_number = str(row[1]).strip()
        estate_raw  = str(row[2]).strip() if row[2] else ''
        cat_raw     = str(row[3]).strip() if row[3] else ''
        status_raw  = str(row[4]).strip() if row[4] else ''
        created_raw = row[5]
        title_raw   = str(row[6]).strip() if row[6] else ''
        log_date_raw= row[7]
        log_content = str(row[8]).strip() if row[8] else ''
        exp_done_raw= row[9]
        staff_raw   = str(row[10]).strip() if row[10] else ''

        # Find estate
        estate_code = task_number.split('/')[1].rstrip('0123456789') if '/' in task_number else ''
        estate = estate_map.get(estate_code)
        if not estate:
            for code, e in estate_map.items():
                if code in estate_raw:
                    estate = e
                    break
        if not estate:
            continue

        # Skip if task already exists
        if task_number in task_cache:
            task = task_cache[task_number]
        else:
            existing = Task.query.filter_by(task_number=task_number).first()
            if existing:
                task_cache[task_number] = existing
                task = existing
            else:
                # Find or create category
                cat = None
                for k, v in cat_map.items():
                    if k in cat_raw or cat_raw in k:
                        cat = v
                        break
                if not cat:
                    cat = cat_map.get('其他')

                # Find status
                status_zh = status_aliases.get(status_raw, '跟進中')
                status = status_map.get(status_zh) or status_map.get('跟進中')

                # Parse created date
                created_at = datetime.utcnow()
                if isinstance(created_raw, datetime):
                    created_at = created_raw
                elif isinstance(created_raw, str):
                    for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
                        try:
                            created_at = datetime.strptime(created_raw, fmt)
                            break
                        except ValueError:
                            pass

                # Parse expected done
                expected_done = None
                if isinstance(exp_done_raw, datetime):
                    expected_done = exp_done_raw.date()
                elif isinstance(exp_done_raw, str) and exp_done_raw not in ('N/A', 'n/a', '跟進中', ''):
                    for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
                        try:
                            expected_done = datetime.strptime(exp_done_raw, fmt).date()
                            break
                        except ValueError:
                            pass

                # Find assignee
                assignee = None
                if staff_raw:
                    # Try exact match first
                    for name, u in user_map.items():
                        if staff_raw.strip().lower() == name.lower():
                            assignee = u
                            break

                task = Task(
                    task_number=task_number,
                    estate_id=estate.id,
                    category_id=cat.id if cat else None,
                    status_id=status.id if status else None,
                    title=title_raw,
                    assignee_id=assignee.id if assignee else None,
                    created_by_id=current_user.id,
                    created_at=created_at,
                    expected_done=expected_done,
                    is_active=True,
                )
                # Mark terminal tasks
                if status and status.is_terminal:
                    task.completed_at = created_at

                db.session.add(task)
                db.session.flush()
                task_cache[task_number] = task
                created_count += 1

                # Update sequence
                from models import TaskNumberSequence
                year = created_at.year
                seq_num = int(re.sub(r'\D', '', task_number.split('/')[1])) if '/' in task_number else 0
                seq = TaskNumberSequence.query.filter_by(estate_id=estate.id, year=year).first()
                if not seq:
                    seq = TaskNumberSequence(estate_id=estate.id, year=year, last_seq=seq_num)
                    db.session.add(seq)
                elif seq.last_seq < seq_num:
                    seq.last_seq = seq_num

        # Add log entry
        if log_content:
            log_date = date.today()
            if isinstance(log_date_raw, datetime):
                log_date = log_date_raw.date()
            elif isinstance(log_date_raw, str):
                for fmt in ['%d/%m/%Y', '%Y-%m-%d']:
                    try:
                        log_date = datetime.strptime(log_date_raw, fmt).date()
                        break
                    except ValueError:
                        pass

            # Check duplicate
            dup = TaskLog.query.filter_by(
                task_id=task.id, log_date=log_date, content=log_content
            ).first()
            if not dup:
                log = TaskLog(
                    task_id=task.id,
                    log_date=log_date,
                    content=log_content,
                    author_id=current_user.id,
                )
                db.session.add(log)
                logs_added += 1

    db.session.commit()
    return {'created': created_count, 'logs_added': logs_added}
